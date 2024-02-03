import openpyxl

def costFunction():
    archivo_excel = 'funcion_de_costo.xlsx'
    libro_trabajo = openpyxl.load_workbook(archivo_excel)
    hoja = libro_trabajo.active
    datos = []

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        datos.append(fila)
    
    libro_trabajo.close()
    
    return datos

def heuristicFunction():
    archivo_excel = 'heuristica.xlsx'
    libro_trabajo = openpyxl.load_workbook(archivo_excel)
    hoja = libro_trabajo.active
    datos = []

    for fila in hoja.iter_rows(min_row=1, values_only=True):
        datos.append(fila)
    
    libro_trabajo.close()
    
    return datos

def getGraph():
    graph = {}
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        graph[origen].append(destino)
        
    return graph

def getWeightedGraph():
    graph = {}
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = cost
        graph[origen].append(dic)
        
    return graph

def getHeuristicGraph():
    heuristicTable = {}
    graph = {}
    for node, value in heuristicFunction()[1:]:
        heuristicTable[node] = value
        
    for origen, destino, _ in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = heuristicTable[destino]
        graph[origen].append(dic)
        
    return graph

def getCompleteGraph():
    heuristicTable = {}
    graph = {}
    for node, value in heuristicFunction()[1:]:
        heuristicTable[node] = value
        
    for origen, destino, cost in costFunction()[1:]:
        if origen not in graph:
            graph[origen] = []
        dic = {}
        dic[destino] = [cost,heuristicTable[destino]]
        graph[origen].append(dic)
        
    return graph